import os, sys
import re
import numpy as np
import pytesseract
from PIL import Image, ImageEnhance, ImageFilter
import cv2
import pandas as pd
from openpyxl import Workbook

support_ext = ('.jpg', '.jpeg', '.png')
entries = os.listdir('assets/')
for entry in entries:
    filename, ext = os.path.splitext(entry)
    if ext in support_ext:
      try:
        img_origin_path = f'assets/{entry}'
        img = cv2.imread(img_origin_path)
        img = cv2.resize(img, (0, 0), fx=3, fy=3)
        kernel = np.ones((2, 2), np.uint8)
        img = cv2.dilate(img, kernel, iterations=1)
        img = cv2.erode(img, kernel, iterations=1)
        img_removed_noise_path = f'assets/{filename}_removed_noise.png'
        cv2.imwrite(img_removed_noise_path, img)

        #Enhance the image
        im = Image.open(img_removed_noise_path)
        img = im.filter(ImageFilter.MedianFilter())
        enhancer = ImageEnhance.Contrast(im)
        img = enhancer.enhance(2)
        img_enhanced_path = f'assets/{filename}_enhance.png'
        img.save(img_enhanced_path)

        with Image.open(img_enhanced_path) as im:
          text = pytesseract.image_to_string(im, lang='chi_sim', nice=10)
          os.remove(img_removed_noise_path)
          os.remove(img_enhanced_path)
          lst = text.split('\n')
          wb = Workbook()
          ws = wb.active
          for row in lst:
            row = re.sub(r" +", '=', row)
            row = row.split('=')
            ws.append(row)
          # reg = re.compile('^(1)[0-9]{10}$')
          # row_count = 0
          # total = len(lst)
          # for txt in lst:
          #   if reg.match(txt):
          #     row_count = total - lst.index(txt)
          #     print(row_count)
          #     break
          # wb = Workbook()
          # ws = wb.active
          # cursor_index = total % row_count
          # column_index = 1
          # while cursor_index < total:
          #   start_index = cursor_index
          #   cursor_index += row_count
          #   column_data = lst[start_index:cursor_index]
          #   for row in range(1, len(column_data)):
          #     ws.cell(column=column_index, row=row, value=column_data[row])
          #   # ws.append(column_data)
          #   column_index += 1

          dist_path = f'result/{filename}.xlsx'
          if os.path.exists(dist_path):
            os.remove(dist_path)
          wb.save(dist_path)
      except OSError as e:
        print('系统内部错误: ', str(e))
      except ValueError as e:
        print('值错误: ', str(e))
      except:
        print('未知错误错误: ', str(sys.exc_info()))
    else:
      print('图片格式不支持')